# STEP 1 From Terminal -Make a New Virtual Environment and Activate
# Install openpyxl
""" Commands in Terminal
# mkdir {python-excel100}
# cd {python-excel100}
# pipenv Install
# pipenv shell
# pip install openpyxl
"""

# STEP2 Create an Excel Worksheet and save in the same dir as virtual envirnoment
# Add test data to the Worksheet such as a column of test scores
""" Example
Column C Rows 4 to 23 contain this data:
17, 14, 17, 16, 11, 20, 19, 11, 7, 15, 16, 17, 18, 19, 14, 13, 16, 20, 19, 7
"""

# STEP3 -Make a new python file and save in same dir as virtual envirnoment
"""
testscore.py
"""

# STEP4 -Add the following code to the testscore.py file
"""   ***************************************    """

# Import the openpyxl library
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Load existing spreadsheet
wb = load_workbook('testscores.xlsx')

# create an active worksheet
ws = wb.active

# Use Print function to test the python file
# testscore1 = ws['C6'].value --(uncomment code for testing)
# print(testscore1) -(uncomment code for testing)
# In Terminal ls -to list files, python testcore.py to run program
"""
(pythonexcel100) bash-3.2$ ls
(pythonexcel100) bash-3.2$ python testscore.py
17
"""

# Make an empty python list to keep the test scores
testscores = []

# Make a new python object called "scores"
# Assign to it the worksheet data from columns C4 to C23
scores = ws['C4': 'C23']
# print(scores)  -(uncomment code for testing)
# At this point, the data will be a tuple object
"""
((<Cell 'Sheet1'.C4>,), (<Cell 'Sheet1'.C5>,), (<Cell 'Sheet1'.C6>,),
(<Cell 'Sheet1'.C7>,), (<Cell 'Sheet1'.C8>,), (<Cell 'Sheet1'.C9>,),
(<Cell 'Sheet1'.C10>,), (<Cell 'Sheet1'.C11>,), (<Cell 'Sheet1'.C12>,),
(<Cell 'Sheet1'.C13>,), (<Cell 'Sheet1'.C14>,), (<Cell 'Sheet1'.C15>,),
(<Cell 'Sheet1'.C16>,), (<Cell 'Sheet1'.C17>,), (<Cell 'Sheet1'.C18>,),
(<Cell 'Sheet1'.C19>,), (<Cell 'Sheet1'.C20>,), (<Cell 'Sheet1'.C21>,),
(<Cell 'Sheet1'.C22>,), (<Cell 'Sheet1'.C23>,))
"""

# Values can not be extraced in a tuple ojbect
# So run a double loop to access the worksheet cell data values
# Append the cell data to the empty python list
for x in scores:
    for z in x:
        testscores.append(z.value)
print(testscores)

# Run python methods to get max value, min value and average
# Assign the results to pyton variables
highscore = max(testscores)
lowscore = min(testscores)
listlen = len(testscores)
total = sum(testscores)
avgscore = total // listlen # use // for division without float

# Assign the values from the above variables to cells in the worksheet
ws['C26'] = highscore
ws['C27'] = lowscore
ws['C28'] = avgscore

# Save the worksheet
wb.save('testscores.xlsx')
print("New File Saved")

# Open excel worksheet to see changes 
